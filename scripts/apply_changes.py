"""
apply_changes.py – ביצוע שינויים בפריטי תפילה דרך ממשק ה-CMS.

קורא קובץ JSON שנוצר על ידי generate_changes.py,
מנווט ב-CMS דרך Playwright, מוצא כל textarea לפי itemId,
מקליד את הטקסט החדש בצורה אנושית, ולוחץ "שמור מקטע".

הרצה ראשונה (כניסה עם חשבון Google):
  python scripts/apply_changes.py --changes scripts/changes_YYYYMMDD_HHMMSS.json --login

הרצות הבאות (סשן שמור):
  python scripts/apply_changes.py --changes scripts/changes_YYYYMMDD_HHMMSS.json

שימוש בפרופיל Chrome קיים:
  python scripts/apply_changes.py --changes scripts/changes_YYYYMMDD_HHMMSS.json --use-chrome --google-email "kpj5722@gmail.com"

דרוש: CMS רץ על http://localhost:5001  (הרץ: npm run dev)
"""

import argparse
import json
import random
import shutil
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from playwright.sync_api import sync_playwright, Page, Locator, TimeoutError as PWTimeout

SCRIPT_DIR = Path(__file__).resolve().parent
USER_DATA_DIR = SCRIPT_DIR / ".playwright-session"
CHROME_SESSION_DIR = SCRIPT_DIR / ".chrome-session"
CMS_URL = "http://localhost:5001/toc-translations"

# נתיב ברירת מחדל לפרופיל Chrome על Windows
DEFAULT_CHROME_PROFILE = Path.home() / "AppData/Local/Google/Chrome/User Data"

# עמודה 3 (קטגוריות): כותרת "3. קטגוריה"
COL3_HEADER = "3. קטגוריה"
# עמודה 4 (תפילות): כותרת "4. תפילה"
COL4_HEADER = "4. תפילה"
# עמודה 5 (מקטעים): כותרת "5. מקטע"
COL5_HEADER = "5. מקטע"

SAVE_TIMEOUT_MS = 20_000
NAV_TIMEOUT_MS = 10_000
LOAD_TIMEOUT_MS = 15_000


def _derive_nusach_id(translation_id: str) -> str:
    """מפיק nusachId מתוך translationId, למשל 0-ashkenaz -> ashkenaz."""
    if "-" not in translation_id:
        return translation_id
    return translation_id.split("-", 1)[1]


def _check_server_running(logger: "Logger") -> None:
    """בודק ששרת ה-CMS רץ על localhost:5001. יוצא עם הוראות אם לא."""
    import urllib.request
    try:
        urllib.request.urlopen(CMS_URL, timeout=4)
    except Exception:
        logger.log(
            f"שגיאה: שרת ה-CMS לא רץ על {CMS_URL}\n\n"
            "הרץ את שרת הפיתוח בטרמינל נפרד:\n"
            "  npm run dev\n\n"
            "לאחר שהשרת עלה (תראי 'Local: http://localhost:5001'),\n"
            "הרץ שוב את הסקריפט."
        )
        sys.exit(1)


def _find_chrome_exe() -> Path | None:
    """מחפש את נתיב chrome.exe על Windows."""
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path.home() / r"AppData\Local\Google\Chrome\Application\chrome.exe",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


# ---------------------------------------------------------------------------
# לוג
# ---------------------------------------------------------------------------

class Logger:
    def __init__(self, log_path: Path):
        self._path = log_path
        self._lines: list[str] = []
        log_path.write_text("", encoding="utf-8")

    def log(self, msg: str, *, print_also: bool = True) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        self._lines.append(line)
        if print_also:
            print(line)
        with open(self._path, "a", encoding="utf-8") as f:
            f.write(line + "\n")

    def section(self, title: str) -> None:
        self.log(f"\n{'='*60}\n{title}\n{'='*60}")


def _iter_chrome_profile_dirs(user_data_dir: Path) -> list[Path]:
    """מחזיר רשימת תיקיות פרופיל רלוונטיות תחת Chrome User Data."""
    profiles: list[Path] = []
    default_dir = user_data_dir / "Default"
    if default_dir.is_dir():
        profiles.append(default_dir)

    for profile_dir in sorted(user_data_dir.glob("Profile *")):
        if profile_dir.is_dir():
            profiles.append(profile_dir)

    return profiles


def _profile_contains_email(profile_dir: Path, email: str) -> bool:
    """
    בדיקה היוריסטית אם המייל מופיע בהעדפות של הפרופיל.
    זה מאפשר לבחור את פרופיל Chrome הנכון בלי לקבע Profile 13.
    """
    email = email.lower()
    candidates = [
        profile_dir / "Preferences",
        profile_dir / "Secure Preferences",
    ]

    for candidate in candidates:
        try:
            text = candidate.read_text(encoding="utf-8", errors="ignore").lower()
        except OSError:
            continue
        if email in text:
            return True

    return False


def _resolve_chrome_profile_directory(
    user_data_dir: Path,
    requested_profile: str | None,
    google_email: str | None,
    logger: Logger,
) -> str | None:
    """
    קובע איזה profile-directory להעביר ל-Chrome.
    אפשר לבחור ידנית, או לאתר לפי כתובת Gmail/Google קיימת בפרופיל.
    """
    if requested_profile:
        profile_path = user_data_dir / requested_profile
        if not profile_path.is_dir():
            raise RuntimeError(
                f"תיקיית הפרופיל לא נמצאה תחת Chrome User Data: {requested_profile}"
            )
        logger.log(f"נבחר פרופיל Chrome ידני: {requested_profile}")
        return requested_profile

    if not google_email:
        logger.log("לא נבחר פרופיל Chrome מפורש; Chrome ייפתח עם ברירת המחדל שלו")
        return None

    matches = [
        profile_dir
        for profile_dir in _iter_chrome_profile_dirs(user_data_dir)
        if _profile_contains_email(profile_dir, google_email)
    ]

    if not matches:
        raise RuntimeError(
            f"לא נמצא פרופיל Chrome שמכיל את המייל {google_email!r}. "
            "נסי להעביר --chrome-profile-directory באופן ידני."
        )

    if len(matches) > 1:
        logger.log(
            "נמצאו כמה פרופילים תואמים למייל; משתמש בראשון: "
            + ", ".join(profile_dir.name for profile_dir in matches)
        )

    selected_profile = matches[0].name
    logger.log(f"נבחר פרופיל Chrome עבור {google_email}: {selected_profile}")
    return selected_profile


def _prepare_chrome_session_clone(
    source_user_data_dir: Path,
    source_profile_directory: str,
    logger: Logger,
) -> tuple[Path, str]:
    """
    משכפל את פרופיל Chrome הנבחר לסשן מקומי זמני.

    זה עוקף מצב שבו Chrome לא פותח remote-debugging על ה-User Data הראשי.
    התוצאה היא user-data-dir זמני עם פרופיל יחיד בשם Default.
    """
    source_profile_dir = source_user_data_dir / source_profile_directory
    if not source_profile_dir.is_dir():
        raise RuntimeError(
            f"תיקיית הפרופיל שנבחרה לא נמצאה: {source_profile_dir}"
        )

    if CHROME_SESSION_DIR.exists():
        logger.log(f"מנקה סשן Chrome זמני קודם: {CHROME_SESSION_DIR}")
        shutil.rmtree(CHROME_SESSION_DIR, ignore_errors=False)

    CHROME_SESSION_DIR.mkdir(parents=True, exist_ok=True)

    local_state = source_user_data_dir / "Local State"
    if local_state.exists():
        shutil.copy2(local_state, CHROME_SESSION_DIR / "Local State")

    target_profile_dir = CHROME_SESSION_DIR / "Default"
    logger.log(
        f"מעתיק נתוני התחברות מפרופיל Chrome '{source_profile_directory}' לסשן זמני: {target_profile_dir}"
    )
    target_profile_dir.mkdir(parents=True, exist_ok=True)

    # מעתיקים רק את מה שסביר שנדרש לסשן של ה-CMS על localhost:
    # cookies + local/session storage + IndexedDB של localhost.
    files_to_copy = [
        "Preferences",
        "Secure Preferences",
    ]
    for relative_file in files_to_copy:
        src = source_profile_dir / relative_file
        if src.exists():
            shutil.copy2(src, target_profile_dir / relative_file)
            logger.log(f"  הועתק קובץ: {relative_file}")

    network_dir = source_profile_dir / "Network"
    if network_dir.is_dir():
        target_network_dir = target_profile_dir / "Network"
        target_network_dir.mkdir(parents=True, exist_ok=True)
        for cookie_file in ("Cookies", "Cookies-journal"):
            src = network_dir / cookie_file
            if src.exists():
                shutil.copy2(src, target_network_dir / cookie_file)
                logger.log(f"  הועתק קובץ: Network/{cookie_file}")

    local_storage_dir = source_profile_dir / "Local Storage"
    if local_storage_dir.is_dir():
        shutil.copytree(
            local_storage_dir,
            target_profile_dir / "Local Storage",
            ignore=shutil.ignore_patterns("LOCK", "lockfile", "*.log"),
        )
        logger.log("  הועתקה תיקייה: Local Storage")

    session_storage_dir = source_profile_dir / "Session Storage"
    if session_storage_dir.is_dir():
        shutil.copytree(
            session_storage_dir,
            target_profile_dir / "Session Storage",
            ignore=shutil.ignore_patterns("LOCK", "lockfile", "*.log"),
        )
        logger.log("  הועתקה תיקייה: Session Storage")

    localhost_indexeddb_dir = source_profile_dir / "IndexedDB" / "http_localhost_5001.indexeddb.leveldb"
    if localhost_indexeddb_dir.is_dir():
        target_indexeddb_parent = target_profile_dir / "IndexedDB"
        target_indexeddb_parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(
            localhost_indexeddb_dir,
            target_indexeddb_parent / "http_localhost_5001.indexeddb.leveldb",
            ignore=shutil.ignore_patterns("LOCK", "lockfile", "*.log"),
        )
        logger.log("  הועתקה תיקייה: IndexedDB/http_localhost_5001.indexeddb.leveldb")

    return CHROME_SESSION_DIR, "Default"


# ---------------------------------------------------------------------------
# הקלדה אנושית
# ---------------------------------------------------------------------------

def human_type(textarea: Locator, new_text: str, logger: Logger) -> None:
    """
    ניקוי שדה textarea והקלדת טקסט חדש בצורה אנושית:
    1. גלילה לאלמנט + פוקוס
    2. בחירת כל התוכן ומחיקה
    3. הקלדה תו-תו עם delay אקראי
    4. אימות שהערך נקלט ב-DOM
    """
    textarea.scroll_into_view_if_needed()
    textarea.click()
    time.sleep(0.15)

    # ניקוי: Ctrl+A → Delete
    textarea.press("Control+a")
    time.sleep(0.05)
    textarea.press("Delete")
    time.sleep(0.1)

    current = textarea.input_value()
    if current:
        # ניסיון נוסף עם triple-click
        textarea.click(click_count=3)
        time.sleep(0.05)
        textarea.press("Delete")
        time.sleep(0.1)

    # הקלדה תו-תו
    for char in new_text:
        textarea.type(char, delay=random.randint(25, 85))

    # אימות DOM
    actual = textarea.input_value()
    if actual != new_text:
        raise ValueError(
            f"אימות DOM נכשל!\n"
            f"  צפוי:  {new_text!r}\n"
            f"  בפועל: {actual!r}"
        )


# ---------------------------------------------------------------------------
# ניווט
# ---------------------------------------------------------------------------

def _column_container(page: Page, header_text: str) -> Locator:
    """מחזיר את div-container של עמודה לפי כותרת h4 שלה."""
    return page.locator(f"div:has(h4:text-is('{header_text}'))")


def select_nusach(page: Page, nusach_id: str, logger: Logger) -> None:
    """בחירת נוסח בעמודה 1, למשל ashkenaz."""
    logger.log(f"  בוחר נוסח: {nusach_id}")

    candidates = [
        page.get_by_role("button", name=nusach_id, exact=True).first,
        page.locator(f'button[title="{nusach_id}"]').first,
        page.locator(f"button:text-is('{nusach_id}')").first,
    ]
    for btn in candidates:
        try:
            btn.wait_for(state="visible", timeout=4_000)
            btn.click()
            time.sleep(0.4)
            return
        except PWTimeout:
            continue

    raise RuntimeError(f"לא נמצא כפתור הנוסח '{nusach_id}' בעמודה 1.")


def select_translation(page: Page, translation_id: str, logger: Logger) -> None:
    """בחירת תרגום בעמודה 2, למשל 0-ashkenaz."""
    logger.log(f"  בוחר תרגום: {translation_id}")

    candidates = [
        page.locator("div:has(> h4:text-is('2. תרגום'))").get_by_role(
            "button", name=translation_id, exact=True
        ).first,
        page.get_by_role("button", name=translation_id, exact=True).first,
        page.locator(f'button[title="{translation_id}"]').first,
        page.locator(f"button:text-is('{translation_id}')").first,
    ]
    for btn in candidates:
        try:
            btn.wait_for(state="visible", timeout=4_000)
            btn.click()
            time.sleep(0.4)
            return
        except PWTimeout:
            continue

    raise RuntimeError(f"לא נמצא כפתור התרגום '{translation_id}' בעמודה 2.")


def select_category(page: Page, category_name: str, logger: Logger) -> None:
    """בחירת קטגוריה בעמודה 3 לפי שם."""
    logger.log(f"  בוחר קטגוריה: '{category_name}'")
    col3 = _column_container(page, COL3_HEADER)
    btn = col3.get_by_role("button", name=category_name, exact=True).first
    btn.wait_for(state="visible", timeout=NAV_TIMEOUT_MS)
    btn.click()
    time.sleep(0.4)


def select_prayer(page: Page, prayer_name: str, logger: Logger) -> None:
    """בחירת תפילה בעמודה 4 לפי שם, לאחר שנבחרה קטגוריה."""
    logger.log(f"  בוחר תפילה: '{prayer_name}'")
    col4 = _column_container(page, COL4_HEADER)
    btn = col4.get_by_role("button", name=prayer_name, exact=True).first
    btn.wait_for(state="visible", timeout=NAV_TIMEOUT_MS)
    btn.click()
    time.sleep(0.4)


def select_part(page: Page, part_name: str, logger: Logger) -> None:
    """בחירת מקטע בעמודה 5 לפי שם."""
    logger.log(f"  בוחר מקטע: '{part_name}'")
    col5 = _column_container(page, COL5_HEADER)
    btn = col5.get_by_role("button", name=part_name, exact=True)
    btn.wait_for(state="visible", timeout=NAV_TIMEOUT_MS)
    btn.click()
    time.sleep(0.5)


def navigate_to_part(
    page: Page,
    nusach_id: str,
    translation_id: str,
    category_name: str | None,
    prayer_name: str | None,
    part_name: str | None,
    logger: Logger,
    current_nav: dict,
) -> None:
    """
    מנווט ל-part הנכון ב-CMS.
    current_nav: מעקב אחרי הניווט הנוכחי כדי לדלג על שלבים כפולים.
    """
    # בחירת נוסח (עמודה 1) ואז תרגום (עמודה 2)
    if current_nav.get("nusachId") != nusach_id:
        select_nusach(page, nusach_id, logger)
        current_nav["nusachId"] = nusach_id
        current_nav.pop("translationId", None)
        current_nav.pop("category", None)
        current_nav.pop("prayer", None)
        current_nav.pop("part", None)

    if current_nav.get("translationId") != translation_id:
        select_translation(page, translation_id, logger)
        current_nav["translationId"] = translation_id
        current_nav.pop("category", None)
        current_nav.pop("prayer", None)
        current_nav.pop("part", None)

    # בחירת קטגוריה (עמודה 3)
    category_key = f"{translation_id}::{category_name}"
    if category_name and current_nav.get("category") != category_key:
        select_category(page, category_name, logger)
        current_nav["category"] = category_key
        current_nav.pop("prayer", None)
        current_nav.pop("part", None)

    # בחירת תפילה (עמודה 4)
    prayer_key = f"{category_key}::{prayer_name}"
    if prayer_name and current_nav.get("prayer") != prayer_key:
        select_prayer(page, prayer_name, logger)
        current_nav["prayer"] = prayer_key
        current_nav.pop("part", None)

    # בחירת מקטע (עמודה 5)
    part_key = f"{prayer_key}::{part_name}"
    if part_name and current_nav.get("part") != part_key:
        select_part(page, part_name, logger)
        current_nav["part"] = part_key
        _wait_for_items_load(page)


def _wait_for_items_load(page: Page) -> None:
    """המתנה עד שרשימת הפריטים נטענת (הודעת 'טוען...' נעלמת)."""
    try:
        page.wait_for_selector(
            "div.animate-pulse:has-text('טוען')",
            state="hidden",
            timeout=LOAD_TIMEOUT_MS,
        )
    except PWTimeout:
        pass
    time.sleep(0.3)


# ---------------------------------------------------------------------------
# עריכת פריט
# ---------------------------------------------------------------------------

def find_textarea_for_item(page: Page, item_id: str) -> Locator:
    """
    מציאת textarea של פריט לפי itemId.
    ב-PartItemRow מוצג: <span>itemId: {curId} | MIT: ...</span>
    ה-textarea נמצא באותו div-כרטיס.

    הגישה: מוצאים את ה-span עם ה-itemId הספציפי, ואז עולים ב-DOM
    לאב הקרוב ביותר שמכיל textarea (XPath ancestor::).
    כך נמנעים מלהחזיר div-עטיפה חיצוני שמכיל את כל הפריטים.
    """
    span = page.locator(f'span:text("itemId: {item_id} |")').first
    row_div = span.locator('xpath=ancestor::div[.//textarea][1]')
    return row_div.locator("textarea").first


def _item_rows_locator(page: Page) -> Locator:
    """
    לוקייטור לכל שורות הפריטים במקטע.
    מחפש כל span עם "itemId:" ועולה לdiv-האב הקרוב שמכיל textarea.
    כך מקבלים בדיוק N divs (אחד לכל פריט) בסדר DOM, ללא div-עטיפות.
    """
    return page.locator('span:text("itemId:")').locator('xpath=ancestor::div[.//textarea][1]')


def apply_item_change_by_index(
    page: Page,
    row_index: int,
    item_id: str,
    before_val: str,
    after_val: str,
    logger: Logger,
) -> None:
    """
    מעדכן פריט לפי מיקום השורה במקטע (0, 1, 2, ...).
    סדר השורות ב-DOM = סדר הפריטים ב-JSON – כך כל פריט מתעדכן בשורה הנכונה.
    """
    logger.log(f"    עורך פריט {row_index + 1} (itemId {item_id})")

    rows = _item_rows_locator(page)
    row = rows.nth(row_index)
    try:
        row.scroll_into_view_if_needed(timeout=NAV_TIMEOUT_MS)
    except PWTimeout:
        raise RuntimeError(
            f"שורה {row_index + 1} (itemId={item_id}) לא בנוף – המקטע עשוי להכיל פחות פריטים."
        )
    time.sleep(0.15)
    textarea = row.locator("textarea").first
    textarea.wait_for(state="visible", timeout=NAV_TIMEOUT_MS)

    current = textarea.input_value()
    if current != before_val:
        preview = (repr(current))[:60]
        logger.log(f"    [אזהרה] ערך קיים שונה מהצפוי – ממשיך: {preview}...")

    human_type(textarea, after_val, logger)
    logger.log(f"    [OK] הקלדה הושלמה (פריט {row_index + 1})")


def apply_item_change(
    page: Page,
    item_id: str,
    field: str,
    before_val: str,
    after_val: str,
    logger: Logger,
) -> None:
    """עדכון שדה בודד בפריט."""
    if field != "content":
        logger.log(f"    [SKIP] שדה '{field}' אינו נתמך (רק content נתמך כרגע)")
        return

    logger.log(f"    עורך item {item_id}: '{before_val}' → '{after_val}'")
    textarea = find_textarea_for_item(page, item_id)

    try:
        textarea.wait_for(state="visible", timeout=NAV_TIMEOUT_MS)
    except PWTimeout:
        raise RuntimeError(
            f"textarea לא נמצא עבור itemId={item_id}. "
            "ודא שהמקטע הנכון טעון."
        )

    # אימות ערך לפני (אזהרה בלבד)
    current = textarea.input_value()
    if current != before_val:
        logger.log(
            f"    [אזהרה] ערך שדה קיים שונה מהצפוי:\n"
            f"      צפוי (before): {before_val!r}\n"
            f"      בפועל:         {current!r}\n"
            f"      ממשיך בכל זאת..."
        )

    human_type(textarea, after_val, logger)
    logger.log(f"    [OK] הקלדה הושלמה ואומתה")


# ---------------------------------------------------------------------------
# שמירה
# ---------------------------------------------------------------------------

def save_part(page: Page, logger: Logger) -> None:
    """לחיצה על כפתור 'שמור מקטע' + המתנה לסיום שמירה."""
    save_btn = page.get_by_role("button", name="שמור מקטע")

    # ודא שהכפתור פעיל (יש שינויים)
    try:
        save_btn.wait_for(state="visible", timeout=5_000)
    except PWTimeout:
        raise RuntimeError("כפתור 'שמור מקטע' לא נמצא בדף.")

    is_disabled = save_btn.is_disabled()
    if is_disabled:
        logger.log("  [אזהרה] כפתור שמור מושבת – ייתכן שאין שינויים שזוהו ב-DOM")
        return

    logger.log("  לוחץ 'שמור מקטע'...")
    save_btn.click()

    # ממתין שהכפתור יחזור למצב מושבת (שמירה הושלמה = אין עוד שינויים)
    try:
        page.wait_for_function(
            "() => { "
            "  const btn = [...document.querySelectorAll('button')]"
            "    .find(b => b.textContent.trim() === 'שמור מקטע');"
            "  return btn && btn.disabled;"
            "}",
            timeout=SAVE_TIMEOUT_MS,
        )
        logger.log("  [OK] שמירה הושלמה")
    except PWTimeout:
        raise RuntimeError(
            "פסק זמן בהמתנה לסיום שמירה. "
            "ודא שה-CMS שמר בהצלחה ובדוק ידנית."
        )

    time.sleep(0.5)


def publish_part(page: Page, logger: Logger) -> None:
    """לחיצה על כפתור '🚀 פרסום (Publish)' + המתנה לאישור פרסום."""
    publish_btn = page.locator("button:has-text('פרסום (Publish)')").first

    try:
        publish_btn.wait_for(state="visible", timeout=5_000)
    except PWTimeout:
        raise RuntimeError("כפתור 'פרסום (Publish)' לא נמצא בדף.")

    if publish_btn.is_disabled():
        logger.log("  [אזהרה] כפתור פרסום מושבת – מדלג")
        return

    logger.log("  לוחץ '🚀 פרסום (Publish)'...")
    publish_btn.click()

    # המתנה קצרה לסיום הפרסום (קריאת API ל-Bagel)
    try:
        page.wait_for_function(
            "() => { "
            "  const btn = [...document.querySelectorAll('button')]"
            "    .find(b => b.textContent.includes('פרסום (Publish)'));"
            "  return btn && !btn.disabled;"
            "}",
            timeout=SAVE_TIMEOUT_MS,
        )
        logger.log("  [OK] פרסום הושלם")
    except PWTimeout:
        logger.log("  [אזהרה] לא התקבל אישור פרסום בזמן – ממשיך בכל זאת")

    time.sleep(0.8)


# ---------------------------------------------------------------------------
# דוח Excel
# ---------------------------------------------------------------------------

def save_excel_report(report_rows: list[dict], output_path: Path, logger: "Logger") -> None:
    """שומר דוח שינויים לקובץ Excel בפורמט זהה ל-apply_report_*.xlsx."""
    if not _OPENPYXL_AVAILABLE:
        logger.log("[אזהרה] openpyxl לא מותקן – דוח Excel לא נוצר. הרץ: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "שינויים"
    ws.sheet_view.rightToLeft = True

    headers = ["נוסח", "תרגום", "קטגוריה", "תפילה", "מקטע", "partId", "itemId", "mit_id", "שדה", "לפני", "אחרי", "סטטוס"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F8F", end_color="2F4F8F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_data in report_rows:
        status = row_data.get("status", "")
        ws.append([
            row_data.get("nusachId", ""),
            row_data.get("translationId", ""),
            row_data.get("categoryName", ""),
            row_data.get("prayerName", ""),
            row_data.get("partName", ""),
            row_data.get("partId", ""),
            row_data.get("itemId", ""),
            row_data.get("mit_id", row_data.get("itemId", "")),
            row_data.get("field", ""),
            row_data.get("before", ""),
            row_data.get("after", ""),
            status,
        ])
        if "הצליח" in status:
            row_fill = success_fill
        elif "נכשל" in status:
            row_fill = fail_fill
        else:
            row_fill = skip_fill
        for cell in ws[ws.max_row]:
            cell.fill = row_fill
            cell.alignment = wrap_align

    col_widths = [12, 15, 15, 15, 20, 12, 18, 18, 10, 40, 40, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    wb.save(output_path)
    logger.log(f"  [OK] דוח Excel נשמר: {output_path}")


# ---------------------------------------------------------------------------
# לוגיקה ראשית
# ---------------------------------------------------------------------------

def group_changes(changes: list[dict]) -> dict:
    """
    קיבוץ שינויים לפי (nusachId, translationId, categoryName, prayerId, partId).
    מחזיר: {(nId, tId, cName, pId, partId): {"categoryName": ..., "prayerName": ..., "partName": ..., "items": [...]}}
    """
    groups: dict = defaultdict(lambda: {"categoryName": None, "prayerName": None, "partName": None, "items": []})
    for change in changes:
        nusach_id = change.get("nusachId") or _derive_nusach_id(change["translationId"])
        key = (
            nusach_id,
            change["translationId"],
            change.get("categoryName"),
            change["prayerId"],
            change.get("partId", ""),
        )
        groups[key]["categoryName"] = change.get("categoryName")
        groups[key]["prayerName"] = change.get("prayerName")
        groups[key]["partName"] = change.get("partName")
        groups[key]["items"].append(change)
    return dict(groups)


def apply_all_changes(page: Page, changes: list[dict], logger: Logger) -> tuple[dict, list[dict]]:
    """
    מבצע את כל השינויים ב-CMS.
    מחזיר: (סיכום, שורות_דוח)
      סיכום: {"success": int, "failed": int, "skipped": int}
      שורות_דוח: רשימת dict לדוח Excel
    """
    groups = group_changes(changes)
    total_groups = len(groups)
    current_nav: dict = {}
    stats = {"success": 0, "failed": 0, "skipped": 0}
    report_rows: list[dict] = []

    logger.section(f"מתחיל עיבוד {total_groups} קבוצות")

    for group_idx, ((nusach_id, translation_id, category_name, prayer_id, part_id), group_data) in enumerate(groups.items(), 1):
        category_name = group_data["categoryName"]
        prayer_name = group_data["prayerName"]
        part_name = group_data["partName"]
        items = group_data["items"]

        logger.section(
            f"קבוצה {group_idx}/{total_groups}: "
            f"{nusach_id} / {translation_id} / קטגוריה={category_name or '-'} / תפילה={prayer_name or prayer_id} / מקטע={part_name or part_id}"
        )

        nav_error: str | None = None
        try:
            navigate_to_part(
                page, nusach_id, translation_id, category_name, prayer_name, part_name, logger, current_nav
            )
        except RuntimeError as e:
            logger.log(f"  [שגיאת ניווט] {e}")
            nav_error = str(e)
            stats["failed"] += len(items)
            current_nav.clear()

        if nav_error:
            for change in items:
                item_id = change["itemId"]
                for field, vals in change.get("fields", {}).items():
                    report_rows.append({
                        "nusachId": nusach_id,
                        "translationId": translation_id,
                        "categoryName": category_name or "",
                        "prayerName": prayer_name or prayer_id,
                        "partName": part_name or part_id,
                        "partId": part_id,
                        "itemId": item_id,
                        "mit_id": change.get("mit_id", item_id),
                        "field": field,
                        "before": vals.get("before") or "",
                        "after": vals.get("after") or "",
                        "status": f"✗ נכשל (ניווט): {nav_error}",
                    })
            continue

        # התחלה מראש רשימת הפריטים
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(0.3)

        group_ok = True
        for idx, change in enumerate(items):
            item_id = change["itemId"]
            fields = change.get("fields", {})

            for field, vals in fields.items():
                before_val = vals.get("before") or ""
                after_val = vals.get("after") or ""
                try:
                    apply_item_change(
                        page, item_id, field, str(before_val), str(after_val), logger
                    )
                    if field == "content":
                        stats["success"] += 1
                        status = "✓ הצליח"
                    else:
                        stats["skipped"] += 1
                        status = "⊘ דולג"
                except Exception as e:
                    logger.log(f"    [שגיאה] פריט {idx + 1} (itemId {item_id}), שדה {field}: {e}")
                    stats["failed"] += 1
                    group_ok = False
                    status = f"✗ נכשל: {e}"

                report_rows.append({
                    "nusachId": nusach_id,
                    "translationId": translation_id,
                    "categoryName": category_name or "",
                    "prayerName": prayer_name or prayer_id,
                    "partName": part_name or part_id,
                    "partId": part_id,
                    "itemId": item_id,
                    "mit_id": change.get("mit_id", item_id),
                    "field": field,
                    "before": str(before_val),
                    "after": str(after_val),
                    "status": status,
                })

        # שמירה (גם אם חלק נכשלו – נשמור מה שניתן)
        try:
            save_part(page, logger)
        except Exception as e:
            logger.log(f"  [שגיאת שמירה] {e}")
            group_ok = False
            stats["failed"] += 1

        if not group_ok:
            current_nav.clear()

    # פרסום אחד בסוף – אחרי שכל המקטעים נשמרו
    logger.section("פרסום כל השינויים")
    try:
        publish_part(page, logger)
    except Exception as e:
        logger.log(f"  [שגיאת פרסום] {e}")

    return stats, report_rows


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="מבצע שינויים ב-CMS דרך Playwright לפי קובץ JSON"
    )
    parser.add_argument(
        "--changes",
        type=Path,
        required=True,
        help="נתיב לקובץ JSON של שינויים (מ-generate_changes.py)",
    )
    parser.add_argument(
        "--login",
        action="store_true",
        help="פתח דפדפן גלוי לכניסה ידנית לחשבון (שומר סשן)",
    )
    parser.add_argument(
        "--use-chrome",
        nargs="?",
        const="auto",
        metavar="PROFILE_PATH",
        help=(
            "השתמש בפרופיל Chrome הקיים (עם חשבון Google מחובר).\n"
            "ברירת מחדל: auto (מאתר אוטומטית).\n"
            "או ציינ/י נתיב ידני: --use-chrome \"C:\\Users\\...\\User Data\""
        ),
    )
    parser.add_argument(
        "--chrome-profile-directory",
        help=(
            "שם תיקיית הפרופיל תחת Chrome User Data, למשל Default או Profile 13. "
            "אם לא צוין, אפשר לבחור אוטומטית לפי --google-email."
        ),
    )
    parser.add_argument(
        "--google-email",
        help=(
            "כתובת Gmail/Google לחיפוש אוטומטי של פרופיל Chrome מתאים, "
            "למשל kpj5722@gmail.com"
        ),
    )
    parser.add_argument(
        "--output-stats",
        type=Path,
        default=None,
        metavar="PATH",
        help="שמור סטטיסטיקת ריצה (JSON) לנתיב שצוין (לשימוש ע\"י run_load_test.py)",
    )
    args = parser.parse_args()

    if not args.changes.exists():
        sys.exit(f"שגיאה: קובץ שינויים לא נמצא: {args.changes}")

    with open(args.changes, encoding="utf-8") as f:
        changes_data = json.load(f)

    changes = changes_data.get("changes", [])
    if not changes:
        print("אין שינויים לביצוע.")
        return

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = SCRIPT_DIR / f"apply_log_{timestamp_str}.txt"
    report_path = SCRIPT_DIR / f"apply_report_{timestamp_str}.xlsx"
    logger = Logger(log_path)

    logger.log(f"קובץ שינויים: {args.changes}")
    logger.log(f"סה\"כ שינויים: {len(changes)}")
    logger.log(f"לוג: {log_path}")

    # בדיקה שהCMS רץ לפני פתיחת הדפדפן
    _check_server_running(logger)

    # בחירת פרופיל דפדפן
    browser = None

    if args.use_chrome:
        chrome_src = Path(args.use_chrome) if args.use_chrome != "auto" else DEFAULT_CHROME_PROFILE
        if not chrome_src.exists():
            logger.log(f"שגיאה: תיקיית Chrome לא נמצאה: {chrome_src}")
            sys.exit(1)

        chrome_exe = _find_chrome_exe()
        if not chrome_exe:
            logger.log("שגיאה: לא נמצא קובץ הפעלה של Chrome. ודא ש-Chrome מותקן.")
            sys.exit(1)

        logger.log(f"משתמש בפרופיל Chrome: {chrome_src}")

        try:
            chrome_profile_directory = _resolve_chrome_profile_directory(
                chrome_src,
                args.chrome_profile_directory,
                args.google_email,
                logger,
            )
        except RuntimeError as e:
            logger.log(f"שגיאה: {e}")
            sys.exit(1)

        if not chrome_profile_directory:
            chrome_profile_directory = "Default"

        # סגירה מלאה של כל תהליכי Chrome (כולל background process)
        logger.log("סוגר תהליכי Chrome קיימים...")
        subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"], capture_output=True)
        time.sleep(2)

        try:
            chrome_runtime_dir, chrome_runtime_profile = _prepare_chrome_session_clone(
                chrome_src, chrome_profile_directory, logger
            )
        except Exception as e:
            logger.log(f"שגיאה בהכנת סשן Chrome זמני: {e}")
            sys.exit(1)
        context_args = dict(
            user_data_dir=str(chrome_runtime_dir),
            executable_path=str(chrome_exe),
            headless=False,
            slow_mo=80,
            args=[
                f"--profile-directory={chrome_runtime_profile}",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-sync",
                "--restore-last-session",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        logger.log("יפתח Chrome ישירות דרך Playwright עם הסשן הזמני")

    else:
        USER_DATA_DIR.mkdir(exist_ok=True)
        context_args = dict(
            user_data_dir=str(USER_DATA_DIR),
            headless=False,
            slow_mo=80,
            args=["--disable-blink-features=AutomationControlled"],
        )

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch_persistent_context(**context_args)
        except Exception as e:
            logger.log(f"שגיאה בפתיחת הדפדפן: {e}")
            sys.exit(1)

        page = browser.pages[0] if browser.pages else browser.new_page()

        time.sleep(1)

        logger.log(f"מנווט ל-{CMS_URL}")
        page.goto(CMS_URL, wait_until="domcontentloaded", timeout=30_000)

        # המתנה לטעינת האפליקציה (React) – מחפש אלמנט מעמודה 1 או כפתור
        try:
            page.wait_for_selector("h4:text-is('1. נוסח'), button:has-text('Sign in')", timeout=15_000)
        except PWTimeout:
            pass
        time.sleep(2)

        # אם צריך כניסה (מופיע כפתור Sign in)
        sign_in_btn = page.locator("button:has-text('Sign in'), a:has-text('Sign in')").first
        if sign_in_btn.count() > 0 or args.login:
            logger.log("נדרשת כניסה – ממתין להתחברות ידנית")
            print("\n" + "="*55)
            print("  הדפדפן נפתח – התחברי עם חשבון Google של קורן.")
            print("  לאחר שנכנסת ל-CMS ורואה את המסך הראשי,")
            print("  חזרי לכאן ולחצי Enter כדי להמשיך.")
            print("="*55)
            input("  לחצי Enter להמשך... ")
            time.sleep(2)
            logger.log("כניסה הושלמה")

        t_start = time.time()
        stats, report_rows = apply_all_changes(page, changes, logger)
        duration_sec = round(time.time() - t_start, 1)
        stats["duration_sec"] = duration_sec

        logger.section(
            f"סיכום:\n"
            f"  הצלחות:  {stats['success']}\n"
            f"  שגיאות:  {stats['failed']}\n"
            f"  דולגו:   {stats['skipped']}\n"
            f"  זמן כולל: {duration_sec} שניות"
        )

        logger.log("שומר דוח Excel...")
        save_excel_report(report_rows, report_path, logger)

        if args.output_stats:
            try:
                args.output_stats.parent.mkdir(parents=True, exist_ok=True)
                with open(args.output_stats, "w", encoding="utf-8") as f:
                    json.dump(stats, f, ensure_ascii=False, indent=2)
                logger.log(f"סטטיסטיקה נשמרה: {args.output_stats}")
            except Exception as e:
                logger.log(f"[אזהרה] שמירת סטטיסטיקה נכשלה: {e}")

        if browser:
            browser.close()

    print(f"\nלוג נשמר:       {log_path}")
    print(f"דוח Excel נשמר: {report_path}")
    if args.output_stats:
        print(f"סטטיסטיקה:      {args.output_stats}")


if __name__ == "__main__":
    main()

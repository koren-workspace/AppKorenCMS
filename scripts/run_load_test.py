"""
run_load_test.py – הרצת בדיקת עומס: ייצור שינויים (הדמיית הקלדה), ביצוע ב-CMS, תיעוד ב-Excel.

שלב 1: מריץ generate_changes עם load_test_config (עד N פריטים).
שלב 2: מריץ apply_changes עם הקובץ שנוצר – הדמיית הקלדה תו-תו + שמירה.
שלב 3: כותב שורה לקובץ Excel עם תאריך, סוג בדיקה, מספר פריטים, הצלחות/שגיאות/דילוג, זמן.

דרוש: pip install openpyxl
הרצה:
  python scripts/run_load_test.py
  python scripts/run_load_test.py --items 100 --report scripts/apply_report.xlsx
  python scripts/run_load_test.py --items 50 --use-chrome --google-email "you@gmail.com"
  python scripts/run_load_test.py --items 100 --concurrent 2 --use-chrome --google-email "you@gmail.com"   # עומס מקבילי – 2 דפדפנים במקביל
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# סביבת UTF-8 לתהליכי apply/generate (חשוב ב-Windows כדי למנוע UnicodeEncodeError)
_ENV_UTF8 = {**os.environ, "PYTHONIOENCODING": "utf-8"}

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG = SCRIPT_DIR / "load_test_config.json"
DEFAULT_REPORT = SCRIPT_DIR / "load_test_report.xlsx"
SHEET_NAME = "עומס"
HEADERS = ["תאריך", "שעה", "סוג בדיקה", "מספר פריטים", "הצלחות", "שגיאות", "דילוג", "זמן (שניות)", "הערות"]


def _ensure_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl import Workbook
        return load_workbook, Workbook
    except ImportError:
        print("נדרש: pip install openpyxl")
        sys.exit(1)


def _latest_changes_json() -> Path | None:
    files = sorted(SCRIPT_DIR.glob("changes_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def _append_report_row(report_path: Path, row: list, load_workbook, Workbook) -> None:
    if report_path.exists():
        wb = load_workbook(report_path)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
    ws.append(row)
    try:
        wb.save(report_path)
    except PermissionError:
        backup_path = report_path.parent / f"{report_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{report_path.suffix}"
        wb.save(backup_path)
        print(f"[אזהרה] הקובץ {report_path.name} נעול (סגרי אותו ב-Excel). נשמר כ: {backup_path.name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="בדיקת עומס: N פריטים, הדמיית הקלדה, תיעוד ב-Excel")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG, help="קונפיג ל-generate_changes")
    parser.add_argument("--items", type=int, default=None, help="מגבלת פריטים (מחליף limit בקונפיג)")
    parser.add_argument(
        "--report",
        type=Path,
        default=DEFAULT_REPORT,
        help=f"קובץ Excel לתיעוד (ברירת מחדל: {DEFAULT_REPORT.name}). לשימוש בקובץ קיים: --report scripts/apply_report.xlsx",
    )
    parser.add_argument("--login", action="store_true", help="פתח דפדפן לכניסה ידנית")
    parser.add_argument("--use-chrome", nargs="?", const="auto", help="שימוש בפרופיל Chrome")
    parser.add_argument("--google-email", help="Gmail לחיפוש פרופיל Chrome")
    parser.add_argument(
        "--runs",
        type=int,
        default=1,
        help="מספר הרצות apply ברצף (כל ריצה מתועדת שורה ב-Excel)",
    )
    parser.add_argument(
        "--concurrent",
        type=int,
        default=1,
        metavar="N",
        help="הרצת N דפדפנים במקביל (עומס מקבילי, אותו חשבון). דורש --use-chrome. כל ריצה עם תיקיית סשן נפרדת.",
    )
    parser.add_argument(
        "--skip-generate",
        action="store_true",
        help="לא להריץ generate_changes – להשתמש בקובץ changes_*.json אחרון",
    )
    args = parser.parse_args()
    if not args.report.is_absolute():
        args.report = SCRIPT_DIR / args.report

    load_workbook, Workbook = _ensure_openpyxl()

    if not args.skip_generate:
        if not args.config.exists():
            sys.exit(f"שגיאה: קובץ config לא נמצא: {args.config}")
        limit_arg = ["--limit", str(args.items)] if args.items else []
        cmd = [sys.executable, str(SCRIPT_DIR / "generate_changes.py"), "--config", str(args.config)] + limit_arg
        print(f"מריץ: {' '.join(cmd)}")
        subprocess.run(cmd, cwd=SCRIPT_DIR.parent, check=True, env=_ENV_UTF8)

    changes_file = _latest_changes_json()
    if not changes_file:
        sys.exit("שגיאה: לא נמצא קובץ changes_*.json. הרץ קודם generate_changes או הסר --skip-generate.")

    print(f"קובץ שינויים: {changes_file}")
    with open(changes_file, encoding="utf-8") as f:
        changes_data = json.load(f)
    changes_list = changes_data.get("changes", [])
    total_items = len(changes_list)
    if total_items == 0:
        sys.exit("אין שינויים בקובץ.")

    apply_cmd = [
        sys.executable,
        str(SCRIPT_DIR / "apply_changes.py"),
        "--changes", str(changes_file),
        "--output-stats", str(SCRIPT_DIR / "load_test_stats.json"),
    ]
    if args.login:
        apply_cmd.append("--login")
    if args.use_chrome:
        apply_cmd.append("--use-chrome")
        if args.use_chrome != "auto":
            apply_cmd.append(args.use_chrome)
    if args.google_email:
        apply_cmd.extend(["--google-email", args.google_email])

    if args.concurrent > 1:
        if not args.use_chrome:
            sys.exit("עומס מקבילי (--concurrent) דורש --use-chrome (אותו חשבון, תיקיות סשן נפרדות).")
        # פיצול השינויים בין ה-workers – כל worker מקבל חתיכה משלו (לא אותו קובץ לכולם)
        n = args.concurrent
        chunk_size = (total_items + n - 1) // n
        chunks = [changes_list[i * chunk_size : (i + 1) * chunk_size] for i in range(n)]
        chunk_files = []
        for i, chunk in enumerate(chunks):
            if not chunk:
                continue
            chunk_data = {**changes_data, "changes": chunk}
            chunk_path = SCRIPT_DIR / f"load_test_chunk_{i + 1}_of_{n}.json"
            with open(chunk_path, "w", encoding="utf-8") as f:
                json.dump(chunk_data, f, ensure_ascii=False, indent=2)
            chunk_files.append((i + 1, chunk_path, len(chunk)))
        # שלב הכנה: סגירת Chrome והעתקת הפרופיל ל-.chrome-session-1, .chrome-session-2 וכו'
        prep_cmd = [
            sys.executable,
            str(SCRIPT_DIR / "apply_changes.py"),
            "--prepare-concurrent-sessions", str(args.concurrent),
            "--use-chrome",
        ]
        if args.use_chrome != "auto":
            prep_cmd.append(args.use_chrome)
        if args.google_email:
            prep_cmd.extend(["--google-email", args.google_email])
        if getattr(args, "chrome_profile_directory", None):
            prep_cmd.extend(["--chrome-profile-directory", args.chrome_profile_directory])
        print("מכין תיקיות סשן לריצה מקבילית (סוגר Chrome, מעתיק פרופיל)...")
        res_prep = subprocess.run(prep_cmd, cwd=SCRIPT_DIR.parent, capture_output=True, text=True, encoding="utf-8", errors="replace", env=_ENV_UTF8)
        if res_prep.returncode != 0:
            print(res_prep.stdout or "")
            print(res_prep.stderr or "", file=sys.stderr)
            sys.exit(f"הכנת סשנים נכשלה (יציאה {res_prep.returncode})")
        from concurrent.futures import ThreadPoolExecutor, as_completed
        def run_apply(suffix: int, changes_path: Path, num_items: int):
            cmd = list(apply_cmd)
            cmd[cmd.index("--changes") + 1] = str(changes_path)
            cmd[cmd.index("--output-stats") + 1] = str(SCRIPT_DIR / f"load_test_stats_concurrent{suffix}.json")
            cmd.extend(["--no-kill-chrome", "--chrome-session-suffix", str(suffix)])
            res = subprocess.run(cmd, cwd=SCRIPT_DIR.parent, capture_output=True, text=True, encoding="utf-8", errors="replace", env=_ENV_UTF8)
            return res, suffix, num_items
        print(f"מריץ {args.concurrent} הרצות apply במקביל (כל אחת עם חתיכת שינויים משלה)...")
        rows_to_add = []
        with ThreadPoolExecutor(max_workers=args.concurrent) as executor:
            futures = [executor.submit(run_apply, suffix, path, num_items) for suffix, path, num_items in chunk_files]
            for fut in as_completed(futures):
                res, suffix, num_items = fut.result()
                stats_path = SCRIPT_DIR / f"load_test_stats_concurrent{suffix}.json"
                if not stats_path.exists():
                    err = (res.stderr or res.stdout or "").strip()
                    note = f"יציאה {res.returncode}"
                    if err:
                        first_line = err.split("\n")[0].strip()[:80]
                        note = f"{note}: {first_line}"
                    rows_to_add.append((
                        datetime.now().strftime("%Y-%m-%d"),
                        datetime.now().strftime("%H:%M:%S"),
                        f"עומס מקביל {suffix}/{args.concurrent}",
                        num_items,
                        "-", "-", "-", "-",
                        note,
                    ))
                else:
                    with open(stats_path, encoding="utf-8") as f:
                        st = json.load(f)
                    rows_to_add.append((
                        datetime.now().strftime("%Y-%m-%d"),
                        datetime.now().strftime("%H:%M:%S"),
                        f"עומס מקביל {suffix}/{args.concurrent}",
                        num_items,
                        st.get("success", 0),
                        st.get("failed", 0),
                        st.get("skipped", 0),
                        st.get("duration_sec", ""),
                        "",
                    ))
        for row in sorted(rows_to_add, key=lambda r: (r[2], r[0], r[1])):
            _append_report_row(args.report, list(row), load_workbook, Workbook)
        for _, path, _ in chunk_files:
            try:
                path.unlink()
            except OSError:
                pass
        print(f"\nתיעוד נשמר ב: {args.report}")
        return

    rows_to_add = []
    for run_idx in range(args.runs):
        if args.runs > 1:
            run_stats_path = SCRIPT_DIR / f"load_test_stats_run{run_idx + 1}.json"
            idx = apply_cmd.index("--output-stats")
            apply_cmd[idx + 1] = str(run_stats_path)
            print(f"מריץ apply_changes ריצה {run_idx + 1}/{args.runs} ({total_items} פריטים)...")
        else:
            print(f"מריץ apply_changes ({total_items} פריטים)...")
        subprocess.run(apply_cmd, cwd=SCRIPT_DIR.parent, check=True, env=_ENV_UTF8)
        current_stats_path = Path(apply_cmd[apply_cmd.index("--output-stats") + 1])
        if not current_stats_path.exists():
            rows_to_add.append((
                datetime.now().strftime("%Y-%m-%d"),
                datetime.now().strftime("%H:%M:%S"),
                f"עומס - הקלדה" + (f" (ריצה {run_idx + 1})" if args.runs > 1 else ""),
                total_items,
                "-", "-", "-", "-",
                "לא נוצר קובץ סטטיסטיקה",
            ))
            continue
        with open(current_stats_path, encoding="utf-8") as f:
            stats = json.load(f)
        rows_to_add.append((
            datetime.now().strftime("%Y-%m-%d"),
            datetime.now().strftime("%H:%M:%S"),
            "עומס - הקלדה" + (f" (ריצה {run_idx + 1}/{args.runs})" if args.runs > 1 else ""),
            total_items,
            stats.get("success", 0),
            stats.get("failed", 0),
            stats.get("skipped", 0),
            stats.get("duration_sec", ""),
            "",
        ))

    for row in rows_to_add:
        _append_report_row(args.report, list(row), load_workbook, Workbook)

    print(f"\nתיעוד נשמר ב: {args.report}")


if __name__ == "__main__":
    main()
